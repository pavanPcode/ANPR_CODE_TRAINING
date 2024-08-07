from ultralytics import YOLO
import cv2
import matplotlib.pyplot as plt
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# Load the models
number_plate_model = YOLO('customYOLO.pt')
character_model = YOLO('best_char_1630.pt')

def plot_results(img, results, model, color=(255, 0, 0)):
    # Loop through results
    for result in results:
        for box in result.boxes.data:
            x1, y1, x2, y2, conf, cls = box
            x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)

            # Draw rectangle around detected object
            cv2.rectangle(img, (x1, y1), (x2, y2), color, 2)

            # Display label
            label = f'{model.names[int(cls)]}:{conf:.2f}'
            cv2.putText(img, label, (x1, y1-10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, color, 2)
    return img

def detect_characters(number_plate_img):
    results = character_model(number_plate_img)
    img = cv2.cvtColor(number_plate_img, cv2.COLOR_BGR2RGB)

    # List to store boxes for sorting
    boxes_list = []

    # Loop through results
    for result in results:
        for box in result.boxes.data:
            x1, y1, x2, y2, conf, cls = box
            boxes_list.append((int(x1), int(y1), int(x2), int(y2), conf, cls))

    # Sort boxes by x1 (left-to-right)
    boxes_list.sort(key=lambda b: b[0])
    plate_text = ""
    # Loop through sorted boxes and draw them
    for x1, y1, x2, y2, conf, cls in boxes_list:
        label = f'{character_model.names[int(cls)]}'
        coordinates = f'({x1}, {y1}), ({x2}, {y2})'
        #print(f'{label} - Coordinates: {coordinates}')
        plate_text += label
    print(f"Vehicle number plate: {plate_text}")
    return plate_text

def process_image(img_path, save_folder):
    # List to store the results for the Excel file
    results_list = []

    # Detect number plate
    number_plate_results = number_plate_model(img_path)

    # Load the image
    img = cv2.imread(img_path)
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

    # Plot number plate results
    img_with_number_plate = plot_results(img, number_plate_results, number_plate_model, color=(255, 0, 0))
    c = 0
    # Loop through the number plate results and extract the number plate image
    for result in number_plate_results:
        plate_text = 'N/A'
        number_plate_img_path = 'N/A'
        for box in result.boxes.data:
            x1, y1, x2, y2, conf, cls = box
            x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
            number_plate_img = img_with_number_plate[y1:y2, x1:x2]
            
            # Detect characters in the number plate
            plate_text = detect_characters(number_plate_img)
            
            # Save the number plate image
            number_plate_img_path = os.path.join(save_folder, f"number_plate_{x1}_{y1}.jpg")
            cv2.imwrite(number_plate_img_path, cv2.cvtColor(number_plate_img, cv2.COLOR_RGB2BGR))
            
        # Append the results to the list
        results_list.append({
            'Image Path': img_path,
            'Number Plate Image Path': number_plate_img_path,
            'Plate Text': plate_text
            })


    # Save the results to an Excel file
    df = pd.DataFrame(results_list)
    excel_path = os.path.join(save_folder, 'number_plate_results.xlsx')
    if os.path.exists(excel_path):
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, index=False, startrow=writer.sheets['Sheet1'].max_row, header=False)
    else:
        df.to_excel(excel_path, index=False)


def process_folder(folder_path):
    # Iterate through all images in the folder
    for filename in os.listdir(folder_path):
        if filename.endswith(('.jpg', '.jpeg', '.png')):
            img_path = os.path.join(folder_path, filename)
            process_image(img_path, save_folder)

    # Add hyperlinks to the Excel file
    excel_path = os.path.join(save_folder, 'number_plate_results.xlsx')
    workbook = load_workbook(excel_path)
    sheet = workbook.active
    for row in range(2, sheet.max_row + 1):
        image_path = sheet.cell(row=row, column=1).value
        number_plate_img_path = sheet.cell(row=row, column=2).value
        if number_plate_img_path and number_plate_img_path != 'N/A':
             # Normalize the path for hyperlink
             number_plate_img_path = number_plate_img_path.replace(os.sep, '/')
             cell = sheet.cell(row=row, column=2)
             cell.hyperlink = f"file:///{number_plate_img_path}"
             cell.font = Font(color="0000FF", underline="single")
        if image_path:
             # Normalize the path for hyperlink
             image_path = image_path.replace(os.sep, '/')
             cell = sheet.cell(row=row, column=1)
             cell.hyperlink = f"file:///{image_path}"
             cell.font = Font(color="0000FF", underline="single")
    workbook.save(excel_path)


# Define the folder to save number plate images
save_folder = r"C:\Users\ADMIN\Desktop\NumberPlate_dataset\Anpr_Code\using_video\numberplates_images"
os.makedirs(save_folder, exist_ok=True)

folder_path = r"C:\Users\ADMIN\Desktop\NumberPlate_dataset\Anpr_Code\using_video\numberplates_images"
process_folder(folder_path)
