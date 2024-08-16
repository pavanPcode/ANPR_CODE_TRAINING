import cv2
import os
from ultralytics import YOLO
from datetime import datetime
import random
import threading
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font

# Create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Vehicle Data"
# Set up headers
worksheet.append(["Vehicle Image", "Number Plate Image", "Number Plate Text"])

# Load the YOLO model
number_plate_model = YOLO('customYOLO.pt')
character_model = YOLO('best_char_1630.pt')

def char_detect(number_plate_img,frame,extracted_count):
    # detect_char
    char_results = character_model(number_plate_img)
    #img = cv2.cvtColor(number_plate_img, cv2.COLOR_BGR2RGB)
    # List to store boxes for sorting
    boxes_list = []
    # Loop through results
    for result in char_results:
        for box in result.boxes.data:
            x1, y1, x2, y2, conf, cls = box
            boxes_list.append((int(x1), int(y1), int(x2), int(y2), conf, cls))
    # Sort boxes by x1 (left-to-right)
    boxes_list.sort(key=lambda b: b[0])
    plate_text = ""
    # Loop through sorted boxes and draw them
    for x1, y1, x2, y2, conf, cls in boxes_list:
        label = f'{character_model.names[int(cls)]}'
        coordinates = f'({x1}, {y1}), ({x2}, {y2})'
        # print(f'{label} - Coordinates: {coordinates}')
        plate_text += label
    print(f"Vehicle number plate: {plate_text}")
    # return plate_text
    # Save the number plate image
    number_plate_img_path = os.path.join(output_folder, f"NP_frame_{extracted_count:04d}_{plate_text}.jpg")
    cv2.imwrite(number_plate_img_path, number_plate_img)

    # save vehicle image
    frame_filename = os.path.join(output_folder, f"VI_frame_{extracted_count:04d}.jpg")
    cv2.imwrite(frame_filename, frame)

    # Normalize the paths for hyperlinks
    number_plate_img_path = number_plate_img_path.replace(os.sep, '/')
    frame_filename = frame_filename.replace(os.sep, '/')

    # Add hyperlinks to the images and insert the detected number plate text
    row = extracted_count + 1  # to account for headers

    # Add hyperlink for the vehicle image
    vehicle_img_cell = worksheet.cell(row=row, column=1)
    vehicle_img_cell.value = "Vehicle Image"
    vehicle_img_cell.hyperlink = f"file:///{frame_filename}"
    vehicle_img_cell.font = Font(color="0000FF", underline="single")

    # Add hyperlink for the number plate image
    number_plate_img_cell = worksheet.cell(row=row, column=2)
    number_plate_img_cell.value = "Number Plate Image"
    number_plate_img_cell.hyperlink = f"file:///{number_plate_img_path}"
    number_plate_img_cell.font = Font(color="0000FF", underline="single")

    # Insert the detected number plate text
    worksheet.cell(row=row, column=3).value = plate_text

    # Save the workbook
    workbook.save(os.path.join(output_folder, "vehicle_data.xlsx"))


def extract_frames(video_path, output_folder, frame_interval=30):
    # Open the video file
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        print("Error: Could not open video.")
        return
    frame_count = 0
    extracted_count = 0

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        # Save every `frame_interval`-th frame
        if frame_count % frame_interval == 0:
            # Detect number plates
            print(datetime.now())
            number_plate_results = number_plate_model(frame)
            c = 0
            for result in number_plate_results:
                print(c,'check')
                for box in result.boxes.data:
                    x1, y1, x2, y2, conf, cls = box
                    x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
                    # Draw the bounding box and label on the frame
                    cv2.rectangle(frame, (x1, y1), (x2, y2), (255, 0, 0), 2)
                    label = f'{number_plate_model.names[int(cls)]}: {conf:.2f}'
                    cv2.putText(frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 0, 0), 2)
                    # Extract the number plate image from the frame
                    number_plate_img = frame[y1:y2, x1:x2]

                    extracted_count += 1
                    #char_detect(number_plate_img, frame, extracted_count)
                    # Create threads with arguments
                    thread1 = threading.Thread(target=char_detect, args=(number_plate_img, frame, extracted_count))
                    # Start threads
                    thread1.start()
                    print(datetime.now(),extracted_count)
                    c = 1
                break
        # Display the frame
        cv2.imshow('Video', frame)
        # Break the loop if 'q' is pressed
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
        frame_count += 1

    cap.release()
    cv2.destroyAllWindows()
    print(f"Extracted {extracted_count} frames from the video.")


# Usage
video_path = r"C:\Users\ADMIN\Desktop\PAVAN\recodeing\license_plate.mp4"
output_folder = r"C:\Users\ADMIN\Desktop\NumberPlate_dataset\Anpr_Code\using_rtsp_test\numberplates_images6"
frame_interval = 60  # Change this to save a different number of frames per second

# Create the output folder if it doesn't exist
if not os.path.exists(output_folder):
    os.makedirs(output_folder)
extract_frames(video_path, output_folder, frame_interval)
