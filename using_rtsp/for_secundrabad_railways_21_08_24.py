import cv2
import os
from ultralytics import YOLO
from datetime import datetime
import threading
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
import time
# Create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Vehicle Data"
# Set up headers
worksheet.append(["Vehicle Image", "Number Plate Image", "Number Plate Text"])

# Load the YOLO model
number_plate_model = YOLO('customYOLO.pt')
character_model = YOLO('best_char_1630.pt')

def char_detect(number_plate_img,frame,extracted_count):
    # detect_char
    char_results = character_model(number_plate_img)
    #img = cv2.cvtColor(number_plate_img, cv2.COLOR_BGR2RGB)
    # List to store boxes for sorting
    boxes_list = []
    # Loop through results
    for result in char_results:
        for box in result.boxes.data:
            x1, y1, x2, y2, conf, cls = box
            boxes_list.append((int(x1), int(y1), int(x2), int(y2), conf, cls))
    # Sort boxes by x1 (left-to-right)
    boxes_list.sort(key=lambda b: b[0])
    plate_text = ""
    # Loop through sorted boxes and draw them
    for x1, y1, x2, y2, conf, cls in boxes_list:
        label = f'{character_model.names[int(cls)]}'
        coordinates = f'({x1}, {y1}), ({x2}, {y2})'
        # print(f'{label} - Coordinates: {coordinates}')
        plate_text += label
    print(f"Vehicle number plate: {plate_text}")
    # return plate_text
    # Save the number plate image
    number_plate_img_path = os.path.join(output_folder, f"NP_frame_{extracted_count:04d}_{plate_text}.jpg")
    cv2.imwrite(number_plate_img_path, number_plate_img)

    # save vehicle image
    frame_filename = os.path.join(output_folder, f"VI_frame_{extracted_count:04d}.jpg")
    cv2.imwrite(frame_filename, frame)

    # Normalize the paths for hyperlinks
    number_plate_img_path = number_plate_img_path.replace(os.sep, '/')
    frame_filename = frame_filename.replace(os.sep, '/')

    # Add hyperlinks to the images and insert the detected number plate text
    row = extracted_count + 1  # to account for headers

    # Add hyperlink for the vehicle image
    vehicle_img_cell = worksheet.cell(row=row, column=1)
    vehicle_img_cell.value = "Vehicle Image"
    vehicle_img_cell.hyperlink = f"file:///{frame_filename}"
    vehicle_img_cell.font = Font(color="0000FF", underline="single")

    # Add hyperlink for the number plate image
    number_plate_img_cell = worksheet.cell(row=row, column=2)
    number_plate_img_cell.value = "Number Plate Image"
    number_plate_img_cell.hyperlink = f"file:///{number_plate_img_path}"
    number_plate_img_cell.font = Font(color="0000FF", underline="single")

    # Insert the detected number plate text
    worksheet.cell(row=row, column=3).value = plate_text

    # Save the workbook
    workbook.save(os.path.join(output_folder, "vehicle_data.xlsx"))

def extract_frames(video_path, output_folder, selected_data,frame_interval=30):
    # Open the video file
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        print("Error: Could not open video.")
        return
    frame_count = 0
    extracted_count = 0

    while True:
        ret, frame = cap.read()
        if not ret:
            break
        cropedframe = frame[selected_data['orig_y1']:selected_data['orig_y2'],
                selected_data['orig_x1']:selected_data['orig_x2']]
        cv2.imshow('Cropped Frame', cropedframe)

        # Resize the frame for display purposes
        screen_width = 1080
        screen_height = 680
        resized_frame = cv2.resize(frame, (screen_width, screen_height))

        # Save every `frame_interval`-th frame
        if frame_count % frame_interval == 0:
            # Detect number plates
            print(datetime.now())
            number_plate_results = number_plate_model(cropedframe,conf=0.3)
            c = 0
            for result in number_plate_results:
                print(c,'check')
                for box in result.boxes.data:
                    x1, y1, x2, y2, conf, cls = box
                    x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
                    # Draw the bounding box and label on the frame
                    cv2.rectangle(resized_frame, (x1, y1), (x2, y2), (255, 0, 0), 2)
                    label = f'{number_plate_model.names[int(cls)]}: {conf:.2f}'
                    cv2.putText(resized_frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 0, 0), 2)
                    # Extract the number plate image from the frame
                    number_plate_img = cropedframe[y1:y2, x1:x2]

                    extracted_count += 1
                    #char_detect(number_plate_img, frame, extracted_count)
                    # Create threads with arguments
                    thread1 = threading.Thread(target=char_detect, args=(number_plate_img, cropedframe, extracted_count))
                    # Start threads
                    thread1.start()
                    print(datetime.now(),extracted_count)
                    c = 1
                break

        # Get the current time with seconds
        current_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())

        # Display the current time on the resized frame
        cv2.putText(resized_frame, current_time, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2, cv2.LINE_AA)

        # Display the frame
        cv2.imshow('Video', resized_frame)
        # Break the loop if 'q' is pressed
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
        frame_count += 1

    cap.release()
    cv2.destroyAllWindows()
    print(f"Extracted {extracted_count} frames from the video.")


def selected_coordinates(video_path):
    # Variables to store rectangle coordinates and control flow
    drawing = [False]  # Use list to allow mutation inside nested function
    ix, iy = [-1], [-1]
    x1, y1, x2, y2 = [0], [0], [0], [0]
    rectangle_drawn = [False]

    # Directory to save the cropped images
    output_dir = "cropped_images"
    os.makedirs(output_dir, exist_ok=True)

    def draw_rectangle(event, x, y, flags, param):
        if event == cv2.EVENT_LBUTTONDOWN and not rectangle_drawn[0]:
            drawing[0] = True
            ix[0], iy[0] = x, y

        elif event == cv2.EVENT_MOUSEMOVE and drawing[0]:
            x1[0], y1[0], x2[0], y2[0] = ix[0], iy[0], x, y

        elif event == cv2.EVENT_LBUTTONUP:
            if drawing[0]:
                drawing[0] = False
                x1[0], y1[0], x2[0], y2[0] = ix[0], iy[0], x, y
                rectangle_drawn[0] = True
                print(f"Rectangle coordinates fixed: ({x1[0]}, {y1[0]}), ({x2[0]}, {y2[0]})")

    # Open the video file
    cap = cv2.VideoCapture(video_path)

    # Check if the video opened successfully
    if not cap.isOpened():
        print("Error: Could not open video.")
        return None, None

    # Read the first frame
    ret, frame = cap.read()
    if not ret:
        print("Error: Could not read the video frame.")
        cap.release()
        return None, None

    # Resize the frame for display purposes
    screen_width = 1080
    screen_height = 680
    resized_frame = cv2.resize(frame, (screen_width, screen_height))

    # Create a window and set the mouse callback function
    cv2.namedWindow('Video')
    cv2.setMouseCallback('Video', draw_rectangle)

    while True:
        # Display the frame with the rectangle (if it's being drawn)
        frame_copy = resized_frame.copy()
        if rectangle_drawn[0]:
            cv2.rectangle(frame_copy, (x1[0], y1[0]), (x2[0], y2[0]), (0, 255, 0), 2)

        cv2.imshow('Video', frame_copy)

        # Press 's' to save the cropped image after drawing the rectangle
        key = cv2.waitKey(1) & 0xFF
        if key == ord('s') and rectangle_drawn[0]:
            # Calculate the original rectangle coordinates before resizing
            orig_x1 = int(x1[0] * (frame.shape[1] / screen_width))
            orig_y1 = int(y1[0] * (frame.shape[0] / screen_height))
            orig_x2 = int(x2[0] * (frame.shape[1] / screen_width))
            orig_y2 = int(y2[0] * (frame.shape[0] / screen_height))

            # Print coordinates for debugging
            print(f"Original rectangle coordinates: ({orig_x1}, {orig_y1}), ({orig_x2}, {orig_y2})")

            # Ensure coordinates are within the frame bounds and valid
            if (orig_x1 >= 0 and orig_y1 >= 0 and orig_x2 <= frame.shape[1] and orig_y2 <= frame.shape[0] and orig_x1 < orig_x2 and orig_y1 < orig_y2):
                # Crop the selected region from the original frame
                cropped_image = frame[orig_y1:orig_y2, orig_x1:orig_x2]

                # Check if cropped image is empty
                if cropped_image.size == 0:
                    print("Error: Cropped image is empty.")
                else:
                    # Save the cropped image
                    save_path = os.path.join(output_dir, "cropped_image.png")
                    cv2.imwrite(save_path, cropped_image)
                    print(f"Cropped image saved at {save_path}")
                break
            else:
                print("Error: Rectangle coordinates are out of frame bounds or invalid.")
        elif key == ord('q'):
            break

    # Release the video capture object and close the display window
    cap.release()
    cv2.destroyAllWindows()
    return {'orig_x1':orig_x1,'orig_y1':orig_y1,'orig_x2':orig_x2, 'orig_y2':orig_y2}


# Usage
video_path = r"C:\Users\ADMIN\Desktop\PAVAN\recodeing\license_plate.mp4"
#video_path = "rtsp://admin:P3r3nni@l@192.168.1.250:554/cam/realmonitor?channel=1&subtype=0"
output_folder = r"C:\Users\ADMIN\Desktop\NumberPlate_dataset\Anpr_Code\using_rtsp_test\numberplates_images3"
frame_interval = 60  # Change this to save a different number of frames per second

# Create the output folder if it doesn't exist
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

selected_data = selected_coordinates(video_path)

extract_frames(video_path, output_folder,selected_data, frame_interval)
