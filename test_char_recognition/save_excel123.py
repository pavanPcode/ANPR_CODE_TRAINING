from ultralytics import YOLO
import cv2
import os
import openpyxl
from openpyxl.utils import get_column_letter

model = YOLO('best.pt')

def plot_and_save_results(img_path, results, input_folder, result_folder):
    # Load img:
    img = cv2.imread(img_path)
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    
    original_img_path = os.path.abspath(img_path)
    result_img_path = os.path.abspath(os.path.join(result_folder, os.path.basename(img_path)))
    
    # List to store boxes for sorting
    boxes_list = []

    # Loop through results
    for result in results:
        for box in result.boxes.data:
            x1, y1, x2, y2, conf, cls = box
            boxes_list.append((int(x1), int(y1), int(x2), int(y2), conf, cls))

    # Sort boxes by y1 (top-to-bottom) and then by x1 (left-to-right)
    boxes_list.sort(key=lambda b: ( b[0]))

    # Loop through results:
    #result_texts = []

    plate_text = ""
    # Loop through sorted boxes and draw them
    for x1, y1, x2, y2, conf, cls in boxes_list:

    #for result in results:
        #for box in result.boxes.data:
            #x1, y1, x2, y2, conf, cls = box
            #x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)

            cv2.rectangle(img, (x1, y1), (x2,y2), (255, 0, 0), 2)

            label = f'{model.names[int(cls)]:{conf:.2f}}'
            plate_text  += label
            #result_texts.append(label)
            cv2.putText(img, label, (x1, y1-10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 0, 0), 2)
            
    # Concatenate all result texts into one string
    #result_text = "\n".join(result_texts)
    
    # Save the result image
    cv2.imwrite(result_img_path, cv2.cvtColor(img, cv2.COLOR_RGB2BGR))
    
    return original_img_path, result_img_path, plate_text

# Define folders
input_folder = r'NumberPlatesForTesting\2\2_input'
result_folder = r'NumberPlatesForTesting\2\2_input-output2'
excel_file = r'NumberPlatesForTesting\2\results20_numbersequence.xlsx'

# Create directories if they don't exist
os.makedirs(result_folder, exist_ok=True)

# Create a new Excel workbook and add a sheet
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = 'Results'

# Add headers
headers = ['Input Image', 'Result Image', 'Result Text']
for col_num, header in enumerate(headers, 1):
    cell = sheet.cell(row=1, column=col_num)
    cell.value = header

# Process all images in the input folder
count = 0
for filename in os.listdir(input_folder):
    img_path = os.path.join(input_folder, filename)
    if os.path.isfile(img_path):
        results = model(img_path)
        original_img_path, result_img_path, result_text = plot_and_save_results(img_path, results, input_folder, result_folder)
        
        count += 1
        
        # Write data to the Excel sheet
        sheet.cell(row=count+1, column=1).value = os.path.basename(original_img_path)
        sheet.cell(row=count+1, column=1).hyperlink = f'file:///{original_img_path.replace("\\", "/")}'
        
        sheet.cell(row=count+1, column=2).value = os.path.basename(result_img_path)
        sheet.cell(row=count+1, column=2).hyperlink = f'file:///{result_img_path.replace("\\", "/")}'
        
        sheet.cell(row=count+1, column=3).value = result_text
        
        print(f'Processed {count} images')

# Save the Excel workbook
workbook.save(excel_file)
print(f'Results saved to {excel_file}')





# from ultralytics import YOLO
# import cv2
# import os
# import csv
# import openpyxl

# model = YOLO('best.pt')

# def plot_and_save_results(img_path, results, input_folder, result_folder):
#     # Load img:
#     img = cv2.imread(img_path)
#     img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    
#     original_img_path = os.path.join(input_folder, os.path.basename(img_path))
#     result_img_path = os.path.join(result_folder, os.path.basename(img_path))
    
#     # Loop through results:
#     result_texts = []
#     for result in results:
#         for box in result.boxes.data:
#             x1, y1, x2, y2, conf, cls = box
#             x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)

#             cv2.rectangle(img, (x1, y1), (x2,y2), (255, 0, 0), 2)

#             label = f'{model.names[int(cls)]:{conf:.2f}}'
#             result_texts.append(label)
#             cv2.putText(img, label, (x1, y1-10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 0, 0), 2)
            
#     # Concatenate all result texts into one string
#     result_text = "".join(result_texts)
#     # Save the result image
#     cv2.imwrite(result_img_path, cv2.cvtColor(img, cv2.COLOR_RGB2BGR))
    
#     return original_img_path, result_img_path, result_text

# # Define folders
# input_folder = r'NumberPlatesForTesting\1\part1-input'
# result_folder = r'NumberPlatesForTesting\1\part1-output'
# csv_file = r'NumberPlatesForTesting\1\results.csv'

# # Create directories if they don't exist
# os.makedirs(result_folder, exist_ok=True)

# # Create and open CSV file
# with open(csv_file, mode='w', newline='') as file:
#     writer = csv.writer(file)
    
#     # Write headers
#     headers = ['Input Image', 'Result Image', 'Result Text']
#     writer.writerow(headers)
    
#     # Process all images in the input folder
#     count = 0
#     for filename in os.listdir(input_folder):
#         img_path = os.path.join(input_folder, filename)
#         if os.path.isfile(img_path):
#             results = model(img_path)
#             original_img_path, result_img_path, result_text = plot_and_save_results(img_path, results, input_folder, result_folder)
            
#             count += 1
            
#             # Write data to the CSV file
#             writer.writerow([original_img_path, result_img_path, result_text])
            
#             print(f'Processed {count} images')
#         break

# print(f'Results saved to {csv_file}')
